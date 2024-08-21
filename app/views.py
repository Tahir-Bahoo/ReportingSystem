from django.shortcuts import render, HttpResponse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from app.models import ExcelFile, Iframe



def index(request):
    
    if request.user.is_anonymous:
        return redirect('home:loginuser')
    
    excel_path = ExcelFile.objects.last()
    excel_input = os.path.abspath(os.getcwd() + f'/media/{excel_path}')

    links_workbook = load_workbook(excel_input, data_only=True)
    sheet = links_workbook['1st - Data Set - Center data']
    header_row = sheet[1]

    header_values = [cell.value for cell in header_row][2::]

    all_centers = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        all_centers.append(row[0])
        # print(row)
    # print(set(all_centers))


    formatted_header_values = []

    for cell in header_values:
        if isinstance(cell, datetime):
            formatted_value = cell.strftime("%b-%Y")
        else:
            formatted_value = cell
        formatted_header_values.append(formatted_value)

    # print(formatted_header_values)
    context = {
        'header_values': formatted_header_values,
        'all_centers': set(all_centers),
    }

    return render(request, 'index.html', context)

def dashboard(request):
    if request.user.is_anonymous:
        return redirect('home:loginuser')
    
    print(Iframe.objects.last().iframe_link)
    
    context = {
        'iframe' : Iframe.objects.last()
    }
    return render(request, 'dashboard.html', context) 




def loginuser(request):
    if request.user.is_authenticated:
        return redirect('home:index')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username = username, password = password)
        if user is not None:
            login(request, user)
            return redirect('home:index')
        else:
            return redirect('home:loginuser')
    
    return render(request, "login.html")



def logoutuser(request):
    logout(request)
    return redirect('home:loginuser')


@csrf_exempt
def generating_report(request):

    if request.method == "POST":
        date = request.POST.get("date")
        center = request.POST.get("center")
        print(date)
        print(center)

        data, all_findings = main(center, date)

        response_data = {
            'all_findings' : all_findings,
        }
        print('\n\n')
        # for key in all_findings:

        #     print(key)           
        #     print(all_findings[key]) 
        #     print('\n\n') 

        for val in data:
            response_data[val] = data[val]
	
        return JsonResponse(response_data)

    return JsonResponse({"messaaage": "error"})


def write(theme, finding, suggestion):
    print(theme,"[[checking]]")
    print(finding, "[[checking]]")
    print(suggestion, "[[checking]]")
#     print('\n\n')
#     print('This is rite Function')
#     print('\n\n')

#     workbook = load_workbook(excel_input)
#     sheet = workbook['4th - Page Result']
#     sheet.append([theme, f'{finding}, {suggestion}'])
#     workbook.save(excel_input)

    


def main(Selected_center, Selected_date):
    
    excel_path = ExcelFile.objects.last()
    excel_input = os.path.abspath(os.getcwd() + f'/media/{excel_path}')

    rows = []
    data = {}
    all_findings = {}
    # config = open('config.txt','r',encoding='utf-8-sig').readlines()
    # config = [x.strip().split('>')[-1] for x in config if x.strip()!='']

    center = str(Selected_center)
    target_date = str(Selected_date)

    # center = 'Centre 3'
    # target_date = 'Oct-2022'

    # center,target_date = config

    links_workbook = load_workbook(excel_input, data_only=True)
    sheet = links_workbook['1st - Data Set - Center data']
    all_center_data = {}

    indx = 0
    for row in sheet.iter_rows(values_only=True):
        # print(row)
        for cell in row:
            if isinstance(cell, datetime):
                cell_value = cell.strftime("%b-%Y")
            else:
                cell_value = str(cell) if cell is not None else ""
            rows.append(cell_value)

        if indx == 0:
            for date in rows[2::]:
                if date == target_date:
                    
                    indx = rows.index(date)
        else:
            if row[0] == center:
                data[row[1]] = row[indx]

            if row[0] == 'All Centers':
                all_center_data[row[1]] = row[indx]

    print(f'>>> Data recorded for {center} on {target_date}')
    # for val in data:
    #     print(f'{val}: {data[val]}')

    print('\n\n')
    print(f'>>> Processing the conditions')
    sheet = links_workbook['2nd - Conditions Script & text']
    for row in sheet.iter_rows(values_only=True):
        try:
            theme = str(row[1]).strip()
            finding = str(row[11]).strip()
            suggestion = str(row[12]).strip()
            parameter_1 = str(row[4]).strip()
            # print('\n\n')
            print(parameter_1, '[Parameter 1]')
            # print('\n\n')

            sign = str(row[5]).strip()
            parameter_2 = str(row[6]).strip()
            try:
                float(parameter_2)
                if float(parameter_2) < 1:
                    finding = finding.replace(
                        '[Parameter 2]', str(float(parameter_2)*100)+'%')
                else:
                    finding = finding.replace(
                        '[Parameter 2]', str(float(parameter_2))+'%')
            except:
                finding = finding.replace('[Parameter 2]', parameter_2)

            finding = finding.replace('[Parameter 1]', parameter_1)
            prm_2_nm = parameter_2[:]
            parameter_1_l = [x for x in data if parameter_1 in x]
            if 'Avg' not in parameter_2:
                parameter_2_l = [x for x in data if parameter_2 in x]
            else:
                parameter_2_l = []

            flag = 1
            if parameter_1_l == []:
                print(
                    f'>>> Parameter 1: {parameter_1} is not there in sheet 1 for {center} on {target_date}')
                flag = 0

            if flag != 0:
                if parameter_2_l == []:
                    if '%' in parameter_2:
                        parameter_2 = float(
                            parameter_2.strip().replace('%', ''))
                        flag = 2
                    else:
                        try:
                            parameter_2 = float(parameter_2.strip())
                            flag = 2
                        except:
                            pass

                    if flag == 1:
                        if 'Avg' not in parameter_2:
                            print(
                                f'>>> Parameter 2: {parameter_2} is not there in sheet 1 for {center} on {target_date}')
                            flag = 0

            if flag > 0:
                parameter_1 = parameter_1_l[0]
                if flag == 1:
                    if 'Avg' not in parameter_2:
                        parameter_2 = parameter_2_l[0]

                prm_1 = data[parameter_1]
                if flag == 1:
                    if 'Avg' in parameter_2:
                        prm_2 = all_center_data[parameter_2]
                    else:
                        prm_2 = data[parameter_2]

                elif flag == 2:
                    prm_2 = parameter_2

                filter_1_match = 0
                if sign == '<':
                    if prm_1 < prm_2:
                        filter_1_match = 1

                elif sign == '>':
                    if prm_1 > prm_2:
                        filter_1_match = 1

                elif sign == '=':
                    if prm_1 == prm_2:
                        filter_1_match = 1

                elif sign == '< =':
                    if prm_1 <= prm_2:
                        filter_1_match = 1

                elif sign == '> =':
                    if prm_1 >= prm_2:
                        filter_1_match = 1

                if filter_1_match == 1:
                    print(
                        f'>>> [FILTER 1 PASSED]: {parameter_1} [{prm_1}] {sign} {prm_2_nm} [{prm_2}]')
                    link = str(row[7]).strip()
                    final_flg = 0

                    if 'AND' in link:
                        parameter_3 = str(row[8]).strip()
                        sign = str(row[9]).strip()
                        parameter_4 = str(row[10]).strip()
                        prm_4_nm = parameter_4

                        parameter_3_l = [
                            x for x in data if parameter_3.lower().strip() in x.lower().strip()]
                        parameter_4_l = [
                            x for x in data if parameter_4.lower().strip() in x.lower().strip()]
                        flag = 1

                        if parameter_3_l == []:
                            print(
                                f'>>> Parameter 3: {parameter_3} is not there in sheet 1 for {center} on {target_date}')
                            flag = 0

                        if flag != 0:
                            if parameter_4_l == []:
                                if '%' in parameter_4:
                                    parameter_4 = float(
                                        parameter_4.strip().replace('%', ''))
                                    flag = 2
                                else:
                                    try:
                                        parameter_4 = float(
                                            parameter_4.strip())
                                        flag = 2
                                    except:
                                        pass

                                if flag == 1:
                                    if 'Avg' not in parameter_4:
                                        print(
                                            f'>>> Parameter 4: {parameter_4} is not there in sheet 1 for {center} on {target_date}')
                                        flag = 0

                        if flag > 0:
                            parameter_3 = parameter_3_l[0]

                            prm_3 = data[parameter_3]
                            if flag == 1:
                                if 'Avg' in parameter_4:
                                    prm_4 = all_center_data[parameter_4]
                                else:
                                    prm_4 = data[parameter_4_l[0]]

                            elif flag == 2:
                                prm_4 = parameter_4

                            filter_2_match = 0
                            if sign == '<':
                                if prm_3 < prm_4:
                                    filter_2_match = 1

                            elif sign == '>':
                                if prm_3 > prm_4:
                                    filter_2_match = 1

                            elif sign == '> =':
                                if prm_3 >= prm_4:
                                    filter_2_match = 1

                            elif sign == '< =':
                                if prm_3 <= prm_4:
                                    filter_2_match = 1

                            elif sign == '=':
                                if prm_3 == prm_4:
                                    filter_2_match = 1

                            if filter_2_match == 1:
                                print(
                                    f'>>> [FILTER 2 PASSED] [AND CONDITION]: {parameter_3} [{prm_3}] {sign} {prm_4_nm} [{prm_4}]')
                                print(
                                    f'>>> [WRITING TO OUTPUT]: AND condition was detected, both filters have passed, writing the text into output')
                                write(theme, finding, suggestion)
                                if theme in all_findings:
                                    all_findings[theme] += "(--)" + finding + "(--)" + suggestion
                                else:
                                    all_findings[theme] = finding + "(--)" + suggestion

                                
                                final_flg = 1
                    else:
                        print(
                            f'>>> [WRITING TO OUTPUT]: Since filter 1 has passed and there is no AND condition so writing the text into output')
                        write(theme, finding, suggestion)
                        if theme in all_findings:
                            all_findings[theme] += "(--)" + finding + "(--)" + suggestion
                        else:
                            all_findings[theme] = finding + "(--)" + suggestion
                else:
                    link = str(row[7]).strip()
                    final_flg = 0
                    if 'OR' in link:
                        parameter_3 = str(row[8]).strip()
                        sign = str(row[9]).strip()
                        parameter_4 = str(row[10]).strip()
                        prm_4_nm = parameter_4
                        parameter_3_l = [
                            x for x in data if parameter_3.lower().strip() in x.lower().strip()]
                        flag = 1

                        if parameter_3_l == []:
                            print(
                                f'>>> Parameter 3: {parameter_3} is not there in sheet 1 for {center} on {target_date}')
                            flag = 0

                        if flag != 0:
                            if '%' in parameter_4:
                                parameter_4 = float(
                                    parameter_4.strip().replace('%', ''))
                                flag = 2
                            else:
                                try:
                                    parameter_4 = float(parameter_4.strip())
                                    flag = 2
                                except:
                                    pass

                            if flag == 1:
                                parameter_4_l = [
                                    x for x in data if parameter_4.lower().strip() in x.lower().strip()]
                                if parameter_4_l == []:
                                    if 'Avg' not in parameter_4:
                                        print(
                                            f'>>> Parameter 4: {parameter_4} is not there in sheet 1 for {center} on {target_date}')
                                        flag = 0

                        if flag > 0:
                            parameter_3 = parameter_3_l[0]

                            prm_3 = data[parameter_3]
                            if flag == 1:
                                if 'Avg' in parameter_4:
                                    prm_4 = all_center_data[parameter_4]
                                else:
                                    prm_4 = data[parameter_4_l[0]]
                            elif flag == 2:
                                prm_4 = parameter_4

                            filter_2_match = 0
                            if sign == '<':
                                if prm_3 < prm_4:
                                    filter_2_match = 1

                            elif sign == '>':
                                if prm_3 > prm_4:
                                    filter_2_match = 1

                            elif sign == '> =':
                                if prm_3 >= prm_4:
                                    filter_2_match = 1

                            elif sign == '< =':
                                if prm_3 <= prm_4:
                                    filter_2_match = 1

                            elif sign == '=':
                                if prm_3 == prm_4:
                                    filter_2_match = 1

                            if filter_2_match == 1:
                                print(
                                    f'>>> [FILTER 2 PASSED] [OR CONDITION]: {parameter_3} [{prm_3}] {sign} {prm_4_nm} [{prm_4}]')
                                print(
                                    f'>>> [WRITING TO OUTPUT]: Since filter 1 did not pass, however there was a OR condition and filter 2 passed so writing the text into output')
                                write(theme, finding, suggestion)
                                if theme in all_findings:
                                    all_findings[theme] += "(--)" + finding + "(--)" + suggestion

                                else:
                                    all_findings[theme] = finding + "(--)" + suggestion
                                final_flg = 1
        except Exception as e:
            print(e, e.__traceback__.tb_lineno, 'learn')


    # print('\n\n')
    # print('all_findings', all_findings)
    # print('\n\n')
    return data, all_findings
